# This Python file uses the following encoding: utf-8
from openpyxl import load_workbook
import json
wb = load_workbook(filename='files/almizan.xlsx', read_only=True)
ws = wb['Sheet1'] # ws is now an IterableWorksheet
lst = {}
counter = 1
for row in ws.rows:
	lst[str(counter)] = []
	lst[str(counter)].append({'title':'المیزان','text':[]})
	for idx in range(0,len(row)):
		if row[idx].value != None:
			lst[str(counter)][0]['text'].append(row[idx].value)
	#print "########>> %s finished"%(str(counter))
	counter +=1

wb = load_workbook(filename='files/nour.xlsx', read_only=True)
ws = wb['Sheet1'] # ws is now an IterableWorksheet
counter = 1
for row in ws.rows:
	if not str(counter) in lst:
		lst[str(counter)] = []
		lst[str(counter)].append({'title':'المیزان','text':[]})
	lst[str(counter)].append({'title':'نور','text':[]})
	for idx in range(0,len(row)):
		if row[idx].value != None:
			lst[str(counter)][1]['text'].append(row[idx].value)
	#print "########>> %s finished"%(str(counter))
	counter +=1

# finalLst = {}
# for lKey in lst:
# 	if len(lst[lKey])==0:
# 		print "%s is empty"%(lKey)
# 	else:
# 		finalLst[lKey] = lst[lKey]

with open('tafsir.json', 'w') as outfile:
	json.dump(lst, outfile)